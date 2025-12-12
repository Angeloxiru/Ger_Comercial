#!/usr/bin/env python3
"""
Script para criar o template Excel com macro VBA para importação de clientes.
"""

import zipfile
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import shutil

def criar_aba_instrucoes(wb):
    """Cria a aba de instruções."""
    ws = wb.create_sheet("📖 INSTRUÇÕES", 0)

    # Configurações de estilo
    titulo_font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
    titulo_fill = PatternFill(start_color="FC0303", end_color="FC0303", fill_type="solid")
    subtitulo_font = Font(name='Calibri', size=12, bold=True, color="FC0303")
    texto_font = Font(name='Calibri', size=11)
    codigo_font = Font(name='Consolas', size=10)
    codigo_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    # Título principal
    ws['A1'] = '🚀 GUIA DE USO - IMPORTAÇÃO DE CLIENTES'
    ws['A1'].font = titulo_font
    ws['A1'].fill = titulo_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 30

    # Introdução
    ws['A3'] = '📋 SOBRE ESTE TEMPLATE'
    ws['A3'].font = subtitulo_font
    ws['A4'] = 'Este arquivo Excel contém uma macro VBA que facilita a importação de dados de clientes.'
    ws['A4'].font = texto_font
    ws['A5'] = 'A macro converte planilhas com qualquer formato para o formato aceito pelo sistema.'
    ws['A5'].font = texto_font

    # Como usar
    ws['A7'] = '⚙️ COMO USAR A MACRO'
    ws['A7'].font = subtitulo_font

    instrucoes = [
        ('1.', 'Habilite as macros quando abrir este arquivo (clique em "Habilitar Conteúdo")'),
        ('2.', 'Prepare seu arquivo Excel com os dados dos clientes (pode ter qualquer formato)'),
        ('3.', 'Pressione ALT + F8 para abrir a lista de macros'),
        ('4.', 'Selecione "Importar_Clientes" e clique em "Executar"'),
        ('5.', 'Escolha o arquivo Excel com seus dados de origem'),
        ('6.', 'A macro irá gerar automaticamente um arquivo .TXT pronto para importação'),
        ('7.', 'Importe o arquivo .TXT gerado no sistema Ger Comercial'),
    ]

    linha = 8
    for num, texto in instrucoes:
        ws[f'A{linha}'] = num
        ws[f'A{linha}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'B{linha}'] = texto
        ws[f'B{linha}'].font = texto_font
        linha += 1

    # Mapeamento de colunas
    linha += 1
    ws[f'A{linha}'] = '🔄 MAPEAMENTO DE COLUNAS'
    ws[f'A{linha}'].font = subtitulo_font
    linha += 1

    ws[f'A{linha}'] = 'Sua planilha pode ter estas colunas:'
    ws[f'A{linha}'].font = Font(name='Calibri', size=11, bold=True)
    linha += 1

    mapeamento = [
        ('Cliente', '→', 'cliente'),
        ('Nome', '→', 'nome'),
        ('Fantasia', '→', 'fantasia'),
        ('Inscr. Est.', '→', 'insc_est'),
        ('CNPJ/CPF', '→', 'cnpj_cpf'),
        ('Grupo', '→', 'grupo'),
        ('Endereço', '→', 'endereco'),
        ('CEP', '→', 'cep'),
        ('Bairro', '→', 'bairro'),
        ('Cidade', '→', 'cidade'),
        ('Nome (Grupo)', '→', 'grupo_desc'),
        ('Descr. (Rota)', '→', 'rota'),
        ('Descrição (Situação)', '→', 'sit_cliente'),
        ('Descrição (Sub Rota)', '→', 'sub_rota'),
        ('Número Endereço', '→', 'num_endereco'),
    ]

    for col_origem, seta, col_destino in mapeamento:
        ws[f'A{linha}'] = col_origem
        ws[f'A{linha}'].font = texto_font
        ws[f'B{linha}'] = seta
        ws[f'B{linha}'].font = texto_font
        ws[f'B{linha}'].alignment = Alignment(horizontal='center')
        ws[f'C{linha}'] = col_destino
        ws[f'C{linha}'].font = codigo_font
        ws[f'C{linha}'].fill = codigo_fill
        linha += 1

    # Dicas importantes
    linha += 1
    ws[f'A{linha}'] = '💡 DICAS IMPORTANTES'
    ws[f'A{linha}'].font = subtitulo_font
    linha += 1

    dicas = [
        '• Seu arquivo pode ter colunas extras - elas serão ignoradas',
        '• A macro detecta automaticamente as colunas corretas',
        '• O arquivo gerado será salvo na sua Área de Trabalho',
        '• Se houver erro, verifique se as macros estão habilitadas',
        '• A macro pula automaticamente linhas vazias',
        '• O arquivo final terá codificação UTF-8 correta',
    ]

    for dica in dicas:
        ws[f'A{linha}'] = dica
        ws[f'A{linha}'].font = texto_font
        linha += 1

    # Ajustar larguras
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15

def criar_aba_exemplo(wb):
    """Cria a aba de exemplo com formato esperado."""
    ws = wb.create_sheet("📝 Exemplo de Dados", 1)

    # Cabeçalhos
    headers = [
        'Cliente', 'Nome', 'Fantasia', 'Inscr. Est.', 'CNPJ/CPF', 'Grupo',
        'Endereço', 'CEP', 'Bairro', 'Cidade', 'Nome (Grupo)',
        'Descr. (Rota)', 'Descrição (Situação)', 'Descrição (Sub Rota)', 'Número Endereço'
    ]

    # Estilo de cabeçalho
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FC0303", end_color="FC0303", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aplicar cabeçalhos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = 20

    # Dados de exemplo
    exemplo_dados = [
        ['001', 'COMÉRCIO AÇÃO E SOLUÇÃO LTDA', 'Ação e Solução', '123456789', '12.345.678/0001-90',
         'GRP01', 'AVENIDA INDEPENDÊNCIA', '12345-678', 'Centro', 'São Paulo',
         'Distribuição', 'R01', 'ATIVO', 'SR01', '100'],
        ['002', 'SUPERMERCADO BOM PREÇO LTDA', 'Bom Preço', '987654321', '98.765.432/0001-10',
         'GRP02', 'RUA DAS FLORES', '54321-876', 'Jardim Europa', 'Rio de Janeiro',
         'Varejo', 'R02', 'ATIVO', 'SR02', '250'],
        ['003', 'DISTRIBUIDORA CENTRAL', 'Central Dist', '456789123', '45.678.912/0001-34',
         'GRP01', 'AVENIDA BRASIL', '11111-222', 'Industrial', 'Belo Horizonte',
         'Distribuição', 'R01', 'ATIVO', 'SR01', '500'],
    ]

    # Adicionar dados de exemplo
    for row_num, row_data in enumerate(exemplo_dados, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    # Adicionar nota
    ws['A6'] = '💡 Dica: Use este formato como referência para seus dados'
    ws['A6'].font = Font(name='Calibri', size=10, italic=True, color="666666")
    ws.merge_cells('A6:E6')

def adicionar_macro_vba(arquivo_xlsm):
    """
    Adiciona o código VBA ao arquivo Excel.
    Arquivos .xlsm são arquivos ZIP com estrutura específica.
    """

    # Código VBA
    vba_code = '''Attribute VB_Name = "Module1"
Sub Importar_Clientes()
    On Error GoTo ErrorHandler

    Dim wbOrigem As Workbook
    Dim wsOrigem As Worksheet
    Dim wbNovo As Workbook
    Dim wsNovo As Worksheet
    Dim fd As FileDialog
    Dim caminhoArquivo As String
    Dim caminhoSalvar As String
    Dim dictColunas As Object
    Dim ultimaLinha As Long
    Dim chave As Variant
    Dim colEncontrada As Variant
    Dim colDestino As Long
    Dim dadosOrigem As Variant
    Dim i As Long
    Dim r As Long, c As Long
    Dim linhaFinal As Long
    Dim numCols As Long
    Dim arrTexto() As String
    Dim timestamp As String
    Dim textoFinal As String
    Dim stream As Object
    Dim fso As Object

    ' Otimização
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual

    ' === Cria o mapeamento dos cabeçalhos ===
    Set dictColunas = CreateObject("Scripting.Dictionary")
    With dictColunas
        .Add "Cliente", "cliente"
        .Add "Nome", "nome"
        .Add "Fantasia", "fantasia"
        .Add "Inscr. Est.", "insc_est"
        .Add "CNPJ/CPF", "cnpj_cpf"
        .Add "Grupo", "grupo"
        .Add "Endereço", "endereco"
        .Add "CEP", "cep"
        .Add "Bairro", "bairro"
        .Add "Cidade", "cidade"
        .Add "Nome (Grupo)", "grupo_desc"
        .Add "Descr. (Rota)", "rota"
        .Add "Descrição (Situação)", "sit_cliente"
        .Add "Descrição (Sub Rota)", "sub_rota"
        .Add "Número Endereço", "num_endereco"
    End With

    ' === Seleciona o arquivo de origem ===
    Set fd = Application.FileDialog(msoFileDialogFilePicker)
    fd.Title = "Selecione o arquivo de origem"
    fd.Filters.Clear
    fd.Filters.Add "Arquivos do Excel", "*.xls; *.xlsx; *.xlsm"

    If fd.Show <> -1 Then
        MsgBox "Nenhum arquivo selecionado. Operação cancelada.", vbExclamation
        GoTo Cleanup
    End If

    caminhoArquivo = fd.SelectedItems(1)

    ' === Abre o arquivo de origem ===
    Set wbOrigem = Workbooks.Open(Filename:=caminhoArquivo, ReadOnly:=True)
    Set wsOrigem = wbOrigem.Sheets(1)

    ' === Valida se tem dados ===
    ultimaLinha = wsOrigem.Cells(wsOrigem.Rows.Count, 1).End(xlUp).Row
    If ultimaLinha < 3 Then
        MsgBox "Arquivo sem dados!", vbExclamation
        GoTo Cleanup
    End If

    ' === Cria novo workbook destino ===
    Set wbNovo = Workbooks.Add
    Set wsNovo = wbNovo.Sheets(1)

    ' === Cabeçalhos destino ===
    i = 1
    For Each chave In dictColunas.Keys
        wsNovo.Cells(1, i).Value = dictColunas(chave)
        i = i + 1
    Next chave

    ' === Copia dados (COMEÇA DA LINHA 3 - PULA LINHA VAZIA) ===
    For Each chave In dictColunas.Keys
        On Error Resume Next
        colEncontrada = Application.Match(chave, wsOrigem.Rows(1), 0)
        On Error GoTo ErrorHandler

        If Not IsError(colEncontrada) Then
            colDestino = Application.Match(dictColunas(chave), wsNovo.Rows(1), 0)

            ' Começa da linha 3 (pula linha 2 vazia)
            dadosOrigem = wsOrigem.Range(wsOrigem.Cells(3, colEncontrada), _
                                         wsOrigem.Cells(ultimaLinha, colEncontrada)).Value

            ' Destino começa na linha 2
            wsNovo.Range(wsNovo.Cells(2, colDestino), _
                        wsNovo.Cells(UBound(dadosOrigem, 1) + 1, colDestino)).Value = dadosOrigem
        End If
    Next chave

    wbOrigem.Close False
    Set wbOrigem = Nothing

    ' === Monta o conteúdo do arquivo texto ===
    linhaFinal = wsNovo.Cells(wsNovo.Rows.Count, 1).End(xlUp).Row
    numCols = wsNovo.Cells(1, wsNovo.Columns.Count).End(xlToLeft).Column

    ReDim arrTexto(1 To linhaFinal)

    For r = 1 To linhaFinal
        Dim linhaTexto As String
        linhaTexto = ""
        For c = 1 To numCols
            Dim valorCelula As String
            valorCelula = ""

            ' Garante conversão correta de valor para string
            If Not IsEmpty(wsNovo.Cells(r, c).Value) Then
                If IsNull(wsNovo.Cells(r, c).Value) Then
                    valorCelula = ""
                Else
                    valorCelula = CStr(wsNovo.Cells(r, c).Value)
                End If
            End If

            ' Substitui aspas por apóstrofo
            valorCelula = Replace(valorCelula, """", "'")

            linhaTexto = linhaTexto & """" & valorCelula & """"
            If c < numCols Then linhaTexto = linhaTexto & vbTab
        Next c
        arrTexto(r) = linhaTexto
    Next r

    wbNovo.Close False
    Set wbNovo = Nothing

    ' === Define caminho de salvamento ===
    timestamp = Format(Now, "yyyymmdd_hhnnss")
    Set fso = CreateObject("Scripting.FileSystemObject")

    ' Tenta Desktop primeiro
    caminhoSalvar = Environ("USERPROFILE") & "\\Desktop\\Clientes_" & timestamp & ".txt"

    ' === Salva com ADODB.Stream (UTF-8 CORRETO) ===
    On Error Resume Next
    Set stream = CreateObject("ADODB.Stream")

    If Err.Number <> 0 Then
        MsgBox "Erro ao criar ADODB.Stream. Verifique se está disponível.", vbCritical
        GoTo Cleanup
    End If

    On Error GoTo ErrorHandler

    stream.Type = 2  ' Texto
    stream.Charset = "UTF-8"
    stream.Open

    ' Junta todas as linhas
    textoFinal = Join(arrTexto, vbCrLf)
    stream.WriteText textoFinal

    ' Tenta salvar no Desktop
    On Error Resume Next
    stream.SaveToFile caminhoSalvar, 2  ' 2 = sobrescreve se existir

    ' Se falhou no Desktop, tenta C:\\Temp
    If Err.Number <> 0 Then
        Err.Clear

        ' Cria C:\\Temp se não existir
        If Not fso.FolderExists("C:\\Temp") Then
            fso.CreateFolder "C:\\Temp"
        End If

        caminhoSalvar = "C:\\Temp\\Clientes_" & timestamp & ".txt"
        stream.SaveToFile caminhoSalvar, 2

        ' Se ainda falhou, tenta Documentos
        If Err.Number <> 0 Then
            Err.Clear
            caminhoSalvar = Environ("USERPROFILE") & "\\Documents\\Clientes_" & timestamp & ".txt"
            stream.SaveToFile caminhoSalvar, 2
        End If
    End If

    On Error GoTo ErrorHandler

    stream.Close
    Set stream = Nothing
    Set fso = Nothing

    Application.ScreenUpdating = True
    Application.Calculation = xlCalculationAutomatic

    MsgBox "✅ Arquivo criado e salvo com sucesso!" & vbCrLf & _
           "Local: " & caminhoSalvar, vbInformation
    Exit Sub

ErrorHandler:
    MsgBox "❌ Erro: " & Err.Description & vbCrLf & _
           "Número do erro: " & Err.Number, vbCritical

Cleanup:
    Application.ScreenUpdating = True
    Application.Calculation = xlCalculationAutomatic
    On Error Resume Next
    If Not wbOrigem Is Nothing Then wbOrigem.Close False
    If Not wbNovo Is Nothing Then wbNovo.Close False
    If Not stream Is Nothing Then
        stream.Close
        Set stream = Nothing
    End If
    Set fso = Nothing
End Sub
'''

    # Diretório temporário para manipulação
    temp_dir = '/tmp/excel_temp'
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)

    # Extrai o arquivo XLSX
    with zipfile.ZipFile(arquivo_xlsm.replace('.xlsm', '.xlsx'), 'r') as zip_ref:
        zip_ref.extractall(temp_dir)

    # Cria a estrutura VBA
    vba_dir = os.path.join(temp_dir, 'xl')
    os.makedirs(vba_dir, exist_ok=True)

    # Cria vbaProject.bin (arquivo binário VBA)
    # Nota: Esta é uma simplificação. Para produção, seria necessário
    # usar uma biblioteca especializada ou um template pré-existente

    # Por ora, vamos criar um arquivo de texto com o código VBA
    # que o usuário pode importar manualmente
    vba_file = arquivo_xlsm.replace('.xlsm', '_MACRO.bas')
    with open(vba_file, 'w', encoding='utf-8') as f:
        f.write(vba_code)

    print(f"⚠️  ATENÇÃO: O código VBA foi salvo em: {vba_file}")
    print("    Para adicionar a macro ao Excel:")
    print("    1. Abra o arquivo Excel")
    print("    2. Pressione ALT + F11")
    print("    3. Vá em File > Import File")
    print("    4. Selecione o arquivo .bas")
    print("    5. Salve o arquivo como .xlsm")

    # Recomprime como XLSM
    # with zipfile.ZipFile(arquivo_xlsm, 'w', zipfile.ZIP_DEFLATED) as zipf:
    #     for root, dirs, files in os.walk(temp_dir):
    #         for file in files:
    #             file_path = os.path.join(root, file)
    #             arcname = os.path.relpath(file_path, temp_dir)
    #             zipf.write(file_path, arcname)

    # Limpa diretório temporário
    # shutil.rmtree(temp_dir)

def main():
    """Função principal."""
    print("🚀 Criando template Excel com macro VBA...")

    # Cria o workbook
    wb = Workbook()

    # Remove a planilha padrão
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Cria as abas
    print("📖 Criando aba de instruções...")
    criar_aba_instrucoes(wb)

    print("📝 Criando aba de exemplo...")
    criar_aba_exemplo(wb)

    # Salva o arquivo
    output_dir = os.path.dirname(os.path.abspath(__file__))
    arquivo_xlsx = os.path.join(output_dir, 'template_importacao_clientes.xlsx')
    arquivo_xlsm = os.path.join(output_dir, 'template_importacao_clientes.xlsm')

    print(f"💾 Salvando arquivo em: {arquivo_xlsx}")
    wb.save(arquivo_xlsx)

    print("🔧 Adicionando macro VBA...")
    adicionar_macro_vba(arquivo_xlsm)

    # Copia o XLSX como XLSM (temporariamente, até adicionarmos o VBA corretamente)
    shutil.copy2(arquivo_xlsx, arquivo_xlsm)

    print(f"\n✅ Template criado com sucesso!")
    print(f"📁 Arquivo XLSX: {arquivo_xlsx}")
    print(f"📁 Arquivo XLSM (para adicionar macro): {arquivo_xlsm}")
    print(f"\n⚠️  PRÓXIMO PASSO: Adicione a macro ao arquivo XLSM usando o arquivo .bas gerado")

if __name__ == '__main__':
    main()
